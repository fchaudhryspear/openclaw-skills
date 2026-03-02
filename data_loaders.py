import requests
import openpyxl
import math as _math
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta # Not strictly needed here, but good to keep if used elsewhere

# ============================================================
# HARDCODED BASE CASE CURVES (Ground Truth from Excel)
# These are the source of truth for the standalone calculator
# ============================================================

# BASE_SETTLEMENT_CURVE: Monthly settlement timing (60 months)
# These are the ground truth values from Excel - DO NOT SCALE
# Performance Guarantee affects the HURDLE, not the settlement curve
# At any PG level, total gross fees = $12.58M (62.13% cumulative settlement)
BASE_SETTLEMENT_CURVE = [
    0.0, 0.0, 0.007243, 0.014667, 0.019634, 0.025149, 0.026289, 0.027454, 0.024697, 0.022209,
    0.019887, 0.019482, 0.017374, 0.017038, 0.013409, 0.013189, 0.011347, 0.011176, 0.011004, 0.010833,
    0.009157, 0.009047, 0.008955, 0.008863, 0.008771, 0.008679, 0.008587, 0.008495, 0.008403, 0.008311,
    0.008238, 0.008164, 0.008091, 0.008018, 0.007944, 0.007871, 0.007797, 0.00776, 0.007724, 0.007687,
    0.007668, 0.007668, 0.007668, 0.007668, 0.007668, 0.007157, 0.007157, 0.007157, 0.007157, 0.007157,
    0.007157, 0.007157, 0.007157, 0.006646, 0.006646, 0.006646, 0.006646, 0.005112, 0.005112, 0.005112, 0.005112
]

# BASE_SURVIVAL_CURVE: Portfolio survival rate (60 months)
# Represents % of enrolled debt still active each month
BASE_SURVIVAL_CURVE = [
    1.0, 1.0, 0.9657, 0.9167, 0.8726, 0.8383, 0.8089, 0.7844, 0.7599, 0.7403,
    0.7232, 0.7085, 0.695, 0.6815, 0.6705, 0.6595, 0.6484, 0.6386, 0.6288, 0.619,
    0.6105, 0.6031, 0.597, 0.5909, 0.5847, 0.5786, 0.5725, 0.5664, 0.5602, 0.5541,
    0.5492, 0.5443, 0.5394, 0.5345, 0.5296, 0.5247, 0.5198, 0.5174, 0.5149, 0.5125,
    0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112,
    0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112, 0.5112
]

# Legacy BORROWER_BASE_CAP_PCT kept for backward compatibility (not used in new model)
BORROWER_BASE_CAP_PCT = [
    1.0000, 1.0000, 1.0000, 0.9585, 0.8952, 0.8047, 0.7143, 0.6238, 0.5334, 0.4429,
    0.3525, 0.2620, 0.1716, 0.0811, 0.0676, 0.0541, 0.0406, 0.0270, 0.0135, 0.0000,
    0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000,
    0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000,
    0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000,
    0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000
]

# Legacy arrays for backwards compatibility
CANCELLATION_DISTRIBUTION = [
    0.0025, 0.07, 0.1, 0.09, 0.07, 0.06, 0.05, 0.05, 0.04, 0.035,
    0.03, 0.0275, 0.0275, 0.0225, 0.0225, 0.0225, 0.02, 0.02, 0.02, 0.0175,
    0.015, 0.0125, 0.0125, 0.0125, 0.0125, 0.0125, 0.0125, 0.0125, 0.0125, 0.01,
    0.01, 0.01, 0.01, 0.01, 0.01, 0.01, 0.005, 0.005, 0.005, 0.0025,
    0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025,
    0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025, 0.0025
]

SETTLEMENT_RATE_CURVE = BASE_SETTLEMENT_CURVE.copy()

# ============================================================
# PERFORMANCE MINIMUM SCHEDULE (Cash Basis) - From Screenshot 1
# Min Performance Factor: 90%, Reset Factor: 95%
# ============================================================

PERFORMANCE_MIN_FACTOR = 0.90
RESET_FACTOR = 0.95

# Measurement periods (months)
MEASUREMENT_PERIODS = [6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 51, 54, 57, 60]

# Targeted Cash Collection % (at each measurement period)
TARGETED_CASH_COLLECTION_PCT = [
    0.251, 0.452, 0.609, 0.723, 0.813, 0.892, 0.964, 1.033, 1.100, 1.166,
    1.230, 1.292, 1.333, 1.339, 1.368, 1.381, 1.394, 1.404
]

# Min Cash Collection % = Targeted × 90%
MIN_CASH_COLLECTION_PCT = [
    0.2259, 0.4066, 0.5445, 0.6506, 0.7320, 0.8026, 0.8672, 0.9298, 0.9904, 1.0494,
    1.1068, 1.1631, 1.1911, 1.2047, 1.2308, 1.2432, 1.2544, 1.2637
]

# Reset to Baseline % = Targeted × 95%
RESET_TO_BASELINE_PCT = [
    0.238, 0.429, 0.575, 0.687, 0.773, 0.847, 0.915, 0.981, 1.045, 1.108,
    1.168, 1.228, 1.257, 1.272, 1.299, 1.312, 1.324, 1.334
]

# FB Settlement Rate (Flobase Settlement Rate at each measurement period)
FB_SETTLEMENT_RATE = [
    0.0837, 0.1506, 0.2017, 0.2409, 0.2711, 0.2973, 0.3212, 0.3444, 0.3668, 0.3887,
    0.4099, 0.4308, 0.4515, 0.4717, 0.4911, 0.5104, 0.5288, 0.5453, 0.5591
]

# NG Settlement Rate (at each measurement period)
NG_SETTLEMENT_RATE = [
    0.0950, 0.1621, 0.2150, 0.2592, 0.2974, 0.3327, 0.3668, 0.3962, 0.4218, 0.4460,
    0.4683, 0.4879, 0.5047, 0.5189, 0.5305, 0.5395, 0.5457, 0.5491, 0.5500
]

# Delta (+/-) between NG and FB settlement rates
SETTLEMENT_RATE_DELTA = [
    0.0113, 0.0115, 0.0133, 0.0183, 0.0263, 0.0354, 0.0456, 0.0518, 0.0550, 0.0573,
    0.0584, 0.0571, 0.0532, 0.0472, 0.0394, 0.0291, 0.0169, 0.0038, -0.0091
]

# ============================================================
# MIN CASH COLLECTION % TO FB OF DEBT LOOKUP - From Screenshot 2
# Maps collection performance to Flobase share of debt
# ============================================================

# Min Cash Collection % values (lookup keys)
MIN_CASH_COLLECTION_LOOKUP = [
    0.1917, 0.3450, 0.4620, 0.5520, 0.6211, 0.6810, 0.7358, 0.7889, 0.8404, 0.8904,
    0.9391, 0.9869, 1.0343, 1.0807, 1.1250, 1.1608, 1.1909, 1.2180, 1.2406
]

# FB of Debt % (corresponding to each Min Cash Collection %)
FB_OF_DEBT_PCT = [
    0.0158, 0.0285, 0.0381, 0.0455, 0.0512, 0.0562, 0.0607, 0.0651, 0.0693, 0.0735,
    0.0775, 0.0814, 0.0853, 0.0892, 0.0928, 0.0958, 0.0982, 0.1005, 0.1023
]

# ============================================================
# FINANCIAL COVENANTS - From Screenshot 3
# Performance thresholds for Collections and Cancellations
# ============================================================

# Collections Covenants (measurement months and thresholds)
COLLECTIONS_COVENANT_MONTHS = [4, 6, 8, 10, 12, 16, 20, 24, 28, 32, 36]
COLLECTIONS_BASE_CASE = [0.112, 0.251, 0.392, 0.506, 0.605, 0.753, 0.867, 0.964, 1.056, 1.144, 1.230]
COLLECTIONS_2_OF_3_MONTHS = [0.120, 0.240, 0.360, 0.460, 0.550, 0.670, 0.760, 0.830, 0.900, 0.970, 1.030]
COLLECTIONS_SINGLE_VINTAGE_FLOOR = [0.100, 0.230, 0.350, 0.450, 0.530, 0.645, 0.730, 0.800, 0.860, 0.915, 0.970]

# Cancellations Covenants (measurement months and thresholds)
CANCELLATIONS_COVENANT_MONTHS = [3, 5, 7, 9, 11, 13, 17, 21, 25, 29, 33]
CANCELLATIONS_BASE_CASE = [0.083, 0.159, 0.208, 0.245, 0.270, 0.289, 0.304, 0.339, 0.352, 0.363, 0.371]
CANCELLATIONS_2_OF_3_MONTHS = [0.110, 0.210, 0.280, 0.330, 0.360, 0.400, 0.410, 0.450, 0.470, 0.480, 0.495]
CANCELLATIONS_SINGLE_VINTAGE_FLOOR = [0.120, 0.230, 0.300, 0.350, 0.390, 0.450, 0.440, 0.460, 0.490, 0.500, 0.520]

# ============================================================
# CUMULATIVE SETTLEMENT RATE CURVES - From Screenshot 4
# Monthly settlement rates for FB and NG scenarios
# ============================================================

# Cumulative Settlement Rate (FB - Flobase) at each month (0-60)
CUMULATIVE_FB_SETTLEMENT_RATE = [
    0.0000, 0.0000, 0.0072, 0.0219, 0.0416, 0.0667, 0.0930, 0.1205, 0.1452, 0.1674,
    0.1873, 0.2068, 0.2242, 0.2412, 0.2553, 0.2687, 0.2800, 0.2912, 0.3022, 0.3131,
    0.3223, 0.3313, 0.3403, 0.3491, 0.3579, 0.3665, 0.3751, 0.3836, 0.3920, 0.4003,
    0.4086, 0.4167, 0.4249, 0.4329, 0.4408, 0.4487, 0.4565, 0.4643, 0.4720, 0.4796,
    0.4873, 0.4950, 0.5027, 0.5103, 0.5180, 0.5252, 0.5323, 0.5395, 0.5466, 0.5538,
    0.5609, 0.5681, 0.5752, 0.5819, 0.5885, 0.5952, 0.6018, 0.6069, 0.6120, 0.6172, 0.6223
]

# Cumulative Settlement Rate (NG - National Guarantee) at each month (0-60)
CUMULATIVE_NG_SETTLEMENT_RATE = [
    0.0000, 0.0000, 0.0082, 0.0248, 0.0471, 0.0755, 0.1053, 0.1365, 0.1645, 0.1896,
    0.2121, 0.2342, 0.2540, 0.2732, 0.2893, 0.3046, 0.3175, 0.3302, 0.3426, 0.3549,
    0.3654, 0.3756, 0.3858, 0.3958, 0.4057, 0.4155, 0.4252, 0.4348, 0.4443, 0.4537,
    0.4631, 0.4723, 0.4814, 0.4905, 0.4994, 0.5083, 0.5171, 0.5258, 0.5345, 0.5431,
    0.5516, 0.5601, 0.5686, 0.5770, 0.5854, 0.5932, 0.6010, 0.6088, 0.6165, 0.6243,
    0.6320, 0.6398, 0.6475, 0.6546, 0.6618, 0.6689, 0.6760, 0.6812, 0.6864, 0.6916, 0.6968
]

def render_news_ticker(sofr_rate, sofr_date):
    """Render a scrolling news ticker with financial market updates including live SOFR"""
    # This function should ideally be in a UI module, but for now we'll keep it here
    # as it's tightly coupled with fetching SOFR for display.
    # In a fully modular app, it would be in a UI helper and take `sofr_display` as input.
    sofr_display = f"SOFR 90-Day Average: {sofr_rate*100:.2f}% (as of {sofr_date}) | Live from NY Fed"
    news_items = [\
        sofr_display,\
        "Student Loan Market: Private lending volumes up 8% YoY",\
        "Credit Markets: Investment-grade spreads tighten to 95bps",\
        "Consumer Credit: Delinquency rates stable at 2.1%",\
        "ABS Market: Student loan securitization issuance reaches $12B in Q4",\
        "Economic Data: Unemployment steady at 3.9%, inflation at 2.4%",\
        "Lending Trends: Average student loan balance rises to $37,850",\
        "Market Update: 10-Year Treasury yield at 4.15%"\
    ]
    ticker_text = "  •  ".join(news_items) + "  •  " + news_items[0]
    
    # In a real Streamlit app, you would use st.markdown here.
    # For a modular file, we just define the logic, not the UI rendering itself.
    # For the purpose of this file, we'll keep the raw markdown for now,
    # but in a production refactor, the `st.markdown` calls would remain in app.py
    # and this function would return the styled HTML string.
    # For now, it's a placeholder.
    ticker_html = f"""
    <style>
    .ticker-wrap {{
        width: 100%;
        overflow: hidden;
        background: linear-gradient(90deg, #1a1a2e 0%, #16213e 50%, #1a1a2e 100%);
        padding: 10px 0;
        border-radius: 5px;
        margin-bottom: 20px;
    }}
    .ticker {{
        display: inline-block;
        white-space: nowrap;
        animation: ticker 45s linear infinite;
        color: #00d4ff;
        font-size: 14px;
        font-weight: 500;
    }}
    .ticker:hover {{
        animation-play-state: paused;
    }}
    @keyframes ticker {{
        0% {{ transform: translateX(0); }}
        100% {{ transform: translateX(-50%); }}
    }}
    </style>
    <div class="ticker-wrap">
        <div class="ticker">{ticker_text} {ticker_text}</div>
    </div>
    """
    return ticker_html # Return HTML to be rendered by Streamlit in app.py

def fetch_sofr_rate_fresh():
    """Fetch the latest SOFR rate from NY Fed - always fresh on first dashboard load"""
    try:
        url = "https://markets.newyorkfed.org/read?productCode=50&eventCodes=525&limit=5&startPosition=0&sort=postDt:-1&format=json"
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if 'refRates' in data and len(data['refRates']) > 0:
                for rate in data['refRates']:
                    if 'average90day' in rate:
                        sofr_90day = float(rate['average90day']) / 100
                        rate_date = rate.get('effectiveDate', 'Unknown')
                        return sofr_90day, rate_date
        return 0.0428, "12/31/2025" # Fallback if API fails
    except Exception as e:
        print(f"Error fetching SOFR: {e}") # Use proper logging in production
        return 0.0428, "12/31/2025" # Fallback

def _log_interpolate(x, x1, x2, y1, y2):
    """Logarithmic interpolation for smooth exponential-like decay"""
    if y1 <= 0 or y2 <= 0:
        return y1 + (x - x1) / (x2 - x1) * (y2 - y1)
    return _math.exp(_math.log(y1) + (x - x1) / (x2 - x1) * (_math.log(y2) - _math.log(y1)))

def build_eop_active_curve(enrolled_debt_base, vintage_term):
    """Build the EOP Active Enrolled Debt curve from known Excel data points"""
    known_pct = {
        0: 1.0, 1: 1.0,
        2: 71_884_294 / 75_000_000,   # 95.85%
        11: 39_265_837 / 75_000_000,  # 52.35% (derived from principal payment)
        12: 37_215_911 / 75_000_000,  # 49.62%
        35: 13_340_493 / 75_000_000,  # 17.79% (derived from principal payment)
        36: 12_631_140 / 75_000_000,  # 16.84%
        37: 0.0,  # Drops to 0 to force full debt payoff
        60: 0.0,
    }
    
    curve = []
    sorted_months = sorted(known_pct.keys())
    
    for m in range(vintage_term + 1):
        if m in known_pct:
            curve.append(known_pct[m] * enrolled_debt_base)
        else:
            prev_months = [k for k in sorted_months if k < m]
            next_months = [k for k in sorted_months if k > m]
            if not prev_months:
                curve.append(known_pct[sorted_months[0]] * enrolled_debt_base)
            elif not next_months:
                curve.append(known_pct[sorted_months[-1]] * enrolled_debt_base)
            else:
                prev_m, next_m = max(prev_months), min(next_months)
                pct = _log_interpolate(m, prev_m, next_m, known_pct[prev_m], known_pct[next_m])
                curve.append(pct * enrolled_debt_base)
    
    return curve

def read_excel_curves(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    cancel_dist = []
    settle_curve = []
    
    try:
        ws_fm = wb['Final Model']
        row95 = list(ws_fm.iter_rows(min_row=95, max_row=95))[0]
        for i in range(19, 79):
            val = row95[i].value if len(row95) > i and row95[i].value else 0
            cancel_dist.append(float(val) if val else 0)
    except Exception as e:
        print(f"Error reading cancellation distribution from Excel: {e}")
        cancel_dist = CANCELLATION_DISTRIBUTION.copy()
    
    try:
        ws_curves = wb['Curvess']
        row3 = list(ws_curves.iter_rows(min_row=3, max_row=3))[0]
        for i in range(5, 66):
            val = row3[i].value if len(row3) > i and row3[i].value else 0
            settle_curve.append(float(val) if val else 0)
    except Exception as e:
        print(f"Error reading settlement curve from Excel: {e}")
        settle_curve = SETTLEMENT_RATE_CURVE.copy()
    
    return cancel_dist, settle_curve

def read_excel_parameters(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb['Final Model']
    
    params = {\
        'advance_rate': ws['D12'].value or 0.08,\
        'preferred_return_factor': ws['D13'].value or 1.3,\
        'pre_preferred_fee_split': ws['D14'].value or 1.0,\
        'post_preferred_fee_split': ws['D15'].value or 0.8,\
        'sofr': ws['D19'].value or 0.0428,\
        'spread': ws['D20'].value or 0.105,\
        'equity_contribution': ws['D21'].value or 0.15,\
        'debt_contribution': ws['D22'].value or 0.85,\
        'loan_term': int(ws['D23'].value or 36),\
        'vintage_term': int(ws['D24'].value or 60),\
        'enrolled_debt': ws['D27'].value or 75000000,\
        'earned_performance_fee': ws['D28'].value or 0.25,\
        'cancellation_rate': ws['D33'].value or 0.49,\
        'egl_accelerant': 0.0,\
    }
    
    cancel_dist, settle_curve = read_excel_curves(uploaded_file)
    params['cancel_dist'] = cancel_dist
    params['settle_curve'] = settle_curve
    
    return params
